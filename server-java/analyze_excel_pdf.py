#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
分析 Excel 和 PDF 文件，进行比对测试
"""

import os
import sys

# 文件路径
BASE_DIR = r"D:\AIProject\PDFTool\原型"
EXCEL_FILE = os.path.join(BASE_DIR, "test.xlsx")
EXPECTED_PDF = os.path.join(BASE_DIR, "test.pdf")
OUTPUT_PDF = os.path.join(BASE_DIR, "my.pdf")

def analyze_excel(file_path):
    """分析 Excel 文件"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, read_only=True)
        print(f"\n{'='*60}")
        print(f"Excel 文件分析: {os.path.basename(file_path)}")
        print(f"{'='*60}")
        print(f"Sheet 数量: {len(wb.sheetnames)}")
        
        for i, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            # 获取使用的范围
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            print(f"  Sheet {i}: '{sheet_name}' - 行数: {max_row}, 列数: {max_col}")
        
        wb.close()
        return len(wb.sheetnames)
    except Exception as e:
        print(f"分析 Excel 失败: {e}")
        return 0

def analyze_pdf(file_path):
    """分析 PDF 文件"""
    try:
        import fitz  # PyMuPDF
        doc = fitz.open(file_path)
        print(f"\n{'='*60}")
        print(f"PDF 文件分析: {os.path.basename(file_path)}")
        print(f"{'='*60}")
        print(f"总页数: {len(doc)}")
        
        page_info = []
        for i, page in enumerate(doc, 1):
            rect = page.rect
            width_mm = rect.width * 25.4 / 72
            height_mm = rect.height * 25.4 / 72
            orientation = "横向" if width_mm > height_mm else "纵向"
            
            # 获取页面文本（部分）
            text = page.get_text()[:100].replace('\n', ' ').strip()
            
            print(f"  第 {i} 页: {width_mm:.1f}mm x {height_mm:.1f}mm ({orientation})")
            print(f"          预览: {text[:50]}..." if len(text) > 50 else f"          预览: {text}")
            
            page_info.append({
                'page': i,
                'width': width_mm,
                'height': height_mm,
                'orientation': orientation,
                'text_preview': text[:100]
            })
        
        doc.close()
        return page_info
    except Exception as e:
        print(f"分析 PDF 失败: {e}")
        return []

def compare_pdfs(expected_info, output_info):
    """比对两个 PDF"""
    print(f"\n{'='*60}")
    print("PDF 比对结果")
    print(f"{'='*60}")
    
    if len(expected_info) != len(output_info):
        print(f"❌ 页数不同: 期望 {len(expected_info)} 页，实际 {len(output_info)} 页")
    else:
        print(f"✓ 页数相同: {len(expected_info)} 页")
    
    min_pages = min(len(expected_info), len(output_info))
    
    for i in range(min_pages):
        exp = expected_info[i]
        out = output_info[i]
        
        print(f"\n第 {i+1} 页比对:")
        
        # 比较方向
        if exp['orientation'] == out['orientation']:
            print(f"  ✓ 方向相同: {exp['orientation']}")
        else:
            print(f"  ❌ 方向不同: 期望 {exp['orientation']}，实际 {out['orientation']}")
        
        # 比较尺寸
        width_diff = abs(exp['width'] - out['width'])
        height_diff = abs(exp['height'] - out['height'])
        
        if width_diff < 5 and height_diff < 5:
            print(f"  ✓ 尺寸接近: {out['width']:.1f}mm x {out['height']:.1f}mm")
        else:
            print(f"  ⚠ 尺寸不同:")
            print(f"    期望: {exp['width']:.1f}mm x {exp['height']:.1f}mm")
            print(f"    实际: {out['width']:.1f}mm x {out['height']:.1f}mm")

def main():
    print("Excel 转 PDF 分析工具")
    print("="*60)
    
    # 检查文件
    if not os.path.exists(EXCEL_FILE):
        print(f"Excel 文件不存在: {EXCEL_FILE}")
        return
    
    if not os.path.exists(EXPECTED_PDF):
        print(f"期望 PDF 不存在: {EXPECTED_PDF}")
        return
    
    # 分析 Excel
    sheet_count = analyze_excel(EXCEL_FILE)
    
    # 分析期望 PDF
    expected_info = analyze_pdf(EXPECTED_PDF)
    
    # 分析输出 PDF（如果存在）
    if os.path.exists(OUTPUT_PDF):
        output_info = analyze_pdf(OUTPUT_PDF)
        compare_pdfs(expected_info, output_info)
    else:
        print(f"\n输出 PDF 不存在: {OUTPUT_PDF}")
        print("请先进行转换测试")
    
    # 输出建议
    print(f"\n{'='*60}")
    print("优化建议")
    print(f"{'='*60}")
    
    if expected_info:
        # 检查期望 PDF 的特征
        orientations = set(p['orientation'] for p in expected_info)
        if len(orientations) == 1:
            print(f"期望 PDF 全部为{list(orientations)[0]}方向")
        else:
            print(f"期望 PDF 包含混合方向: {orientations}")
        
        # 检查页数与 Sheet 数的关系
        if len(expected_info) == sheet_count:
            print(f"✓ 每个 Sheet 正好对应 1 页 PDF (共 {sheet_count} 个)")
        elif len(expected_info) < sheet_count:
            print(f"⚠ PDF 页数 ({len(expected_info)}) 少于 Sheet 数 ({sheet_count})")
            print("  可能有些 Sheet 被合并或跳过了")
        else:
            print(f"⚠ PDF 页数 ({len(expected_info)}) 多于 Sheet 数 ({sheet_count})")
            print("  可能有些 Sheet 被分成多页了")

if __name__ == "__main__":
    main()

